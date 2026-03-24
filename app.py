import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import numpy as np
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Conference Schedule Maker", layout="wide")
st.markdown("""
    <style>
        body { background-color: #121212; color: white; }
        .stRadio>div>label { color: white; }
        .stDataFrame { background-color: #1e1e1e; color: white; }
        .stSelectbox>div>label { color: white; }
        .stNumberInput>div>label { color: white; }
        h1 { text-align: center; margin-bottom: 50px; }
        .column-spacing { padding-left: 70px; padding-right: 70px; }
    </style>
""", unsafe_allow_html=True)

st.markdown("<h1>Conference Schedule Maker</h1>", unsafe_allow_html=True)

# ----------------------------------------------------------------
# HELPERS
# ----------------------------------------------------------------
HEADER_FILL  = PatternFill("solid", start_color="1F3864")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
ROW_FONT     = Font(name="Arial", size=10)
EVEN_FILL    = PatternFill("solid", start_color="DCE6F1")
ODD_FILL     = PatternFill("solid", start_color="FFFFFF")
THIN_BORDER  = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)
COL_WIDTHS   = {"Section": 14, "Date": 12, "Session ID": 11,
                "Time Slot": 11, "Theme": 20, "Title": 40,
                "Presenter(s)": 28, "Faculty Mentor": 24,
                "Availability Constraint": 28,
                "Overflow": 12}

def style_sheet(ws, columns):
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 18)
    ws.row_dimensions[1].height = 20

def write_rows(ws, df, columns, has_date=True):
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL
        for col_idx, col_name in enumerate(columns, start=1):
            val = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = ROW_FONT
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

def make_sheet_name(section_date, start_time, end_time):
    """Format: MMDD HHMM-HHMM  e.g. 415 1430-1600"""
    d = section_date.strftime("%-m%d").lstrip("0") or "0"
    s = start_time.strftime("%H%M")
    e = end_time.strftime("%H%M")
    return f"{d} {s}-{e}"

# ----------------------------------------------------------------
# CONSTRAINT PARSING
# ----------------------------------------------------------------
_WINDOW_RE = re.compile(
    r"(april\s+\d{1,2}),?\s+(\d{1,2}:\d{2})\s*[-\u2013\u2014]\s*(\d{1,2}:\d{2})\s*(am|pm)",
    re.IGNORECASE,
)
_DAY_ONLY_RE = re.compile(r"(?i)^(april\s+\d{1,2})\s+only")

def parse_constraint(raw_value, year):
    """Parse an Availability Constraint cell into a structured dict."""
    if pd.isna(raw_value) or str(raw_value).strip() == "":
        return {"type": "any"}
    s = str(raw_value).strip()
    sl = s.lower()
    if sl.startswith("late submission") or sl == "either day":
        return {"type": "any"}
    if sl.startswith("none") or sl.startswith("neither day"):
        return {"type": "excluded", "note": s}

    m = _DAY_ONLY_RE.match(s)
    if m:
        try:
            d = datetime.strptime(f"{m.group(1)} {year}", "%B %d %Y").date()
            return {"type": "day_only", "dates": [d], "note": s}
        except ValueError:
            return {"type": "unrecognized", "raw": s}

    windows = _WINDOW_RE.findall(s)
    if windows:
        parsed_windows = []
        for date_str, start_str, end_str, ampm in windows:
            try:
                d = datetime.strptime(f"{date_str} {year}", "%B %d %Y").date()
                st = datetime.strptime(f"{start_str} {ampm}", "%I:%M %p").time()
                et = datetime.strptime(f"{end_str} {ampm}", "%I:%M %p").time()
                parsed_windows.append((d, st, et))
            except ValueError:
                continue
        if parsed_windows:
            return {"type": "windows", "windows": parsed_windows}

    return {"type": "unrecognized", "raw": s}


def match_constraint_to_sections(parsed, sections):
    """Return list of section indices the presentation is allowed in."""
    ctype = parsed["type"]
    all_idx = list(range(len(sections)))
    if ctype in ("any", "unrecognized"):
        return all_idx
    if ctype == "excluded":
        return []
    if ctype == "day_only":
        return [i for i, sec in enumerate(sections) if sec["date"] in parsed["dates"]]
    if ctype == "windows":
        matched = []
        for i, sec in enumerate(sections):
            for wd, ws, we in parsed["windows"]:
                if sec["date"] == wd and sec["start"] == ws and sec["end"] == we:
                    matched.append(i)
                    break
                # Fuzzy: same date and overlapping time range
                if sec["date"] == wd:
                    sec_s = datetime.combine(sec["date"], sec["start"])
                    sec_e = datetime.combine(sec["date"], sec["end"])
                    win_s = datetime.combine(wd, ws)
                    win_e = datetime.combine(wd, we)
                    overlap_start = max(sec_s, win_s)
                    overlap_end = min(sec_e, win_e)
                    if overlap_end > overlap_start:
                        matched.append(i)
                        break
        return matched
    return all_idx


def assign_with_constraints(df_work, sections, has_constraints, mode,
                            slot_duration=15, max_presentations=4):
    """
    Constraint-aware scheduling algorithm.
    mode: 'oral' or 'poster'
    Returns (final_df, excluded_df, warnings).
    """
    num_sections = len(sections)
    warnings = []

    if not has_constraints or "Availability Constraint" not in df_work.columns:
        # Fallback: original even-split algorithm
        return _assign_no_constraints(df_work, sections, mode, slot_duration, max_presentations)

    year = sections[0]["date"].year if sections else datetime.today().year

    # Phase 0: classify
    df_work["_parsed"] = df_work["Availability Constraint"].apply(lambda v: parse_constraint(v, year))
    df_work["_allowed"] = df_work["_parsed"].apply(lambda p: match_constraint_to_sections(p, sections))
    df_work["_ctype"] = df_work["_parsed"].apply(lambda p: p["type"])

    excluded_mask = df_work["_ctype"] == "excluded"
    excluded_df = df_work[excluded_mask].copy()

    unrecognized_mask = df_work["_ctype"] == "unrecognized"
    if unrecognized_mask.any():
        warnings.append(f"{unrecognized_mask.sum()} presentation(s) have unrecognized constraint formats (treated as unconstrained).")

    # Presentations that have a constraint referencing a date/time with no matching section
    no_match_mask = (~excluded_mask) & (df_work["_allowed"].apply(len) == 0)
    if no_match_mask.any():
        for _, row in df_work[no_match_mask].iterrows():
            warnings.append(f"'{row['Title']}' has a constraint that doesn't match any configured section — excluded.")
        excluded_df = pd.concat([excluded_df, df_work[no_match_mask]], ignore_index=True)

    schedulable = df_work[~excluded_mask & ~no_match_mask].copy()

    if schedulable.empty:
        df_work["Session ID"] = None
        df_work["Date"] = None
        df_work["Time Slot"] = None
        df_work["Section"] = None
        df_work["Overflow"] = ""
        df_work.drop(columns=["_parsed", "_allowed", "_ctype"], inplace=True, errors="ignore")
        if not excluded_df.empty:
            excluded_df = excluded_df.drop(columns=["_parsed", "_allowed", "_ctype", "Session ID", "Date", "Time Slot", "Section"], errors="ignore")
        return df_work, excluded_df, warnings, 1

    all_idx = set(range(num_sections))
    constrained_mask = schedulable["_allowed"].apply(lambda a: set(a) != all_idx)
    constrained = schedulable[constrained_mask].copy()
    unconstrained = schedulable[~constrained_mask].copy()

    # Phase 1: place constrained (most restricted first)
    if not constrained.empty:
        constrained["_strictness"] = constrained["_allowed"].apply(len)
        constrained = constrained.sort_values(by=["_strictness", "Theme"])

    section_buckets = [[] for _ in range(num_sections)]
    for orig_idx, row in constrained.iterrows():
        allowed = row["_allowed"]
        best = min(allowed, key=lambda i: len(section_buckets[i]))
        section_buckets[best].append(orig_idx)

    # Phase 2: fill unconstrained proportionally
    total_schedulable = len(schedulable)
    target_per_section = total_schedulable / num_sections if num_sections > 0 else 0
    remaining = [max(0, target_per_section - len(bucket)) for bucket in section_buckets]
    total_remaining = sum(remaining)

    unconstrained = unconstrained.sort_values(by="Theme")
    if total_remaining > 0 and len(unconstrained) > 0:
        proportions = [r / total_remaining for r in remaining]
        split_points = [0]
        cumulative = 0
        for p in proportions:
            cumulative += p * len(unconstrained)
            split_points.append(int(round(cumulative)))
        for si in range(num_sections):
            chunk = unconstrained.iloc[split_points[si]:split_points[si + 1]]
            section_buckets[si].extend(chunk.index.tolist())
    elif len(unconstrained) > 0:
        # All sections equally full; just spread evenly
        splits = np.linspace(0, len(unconstrained), num_sections + 1, dtype=int)
        for si in range(num_sections):
            chunk = unconstrained.iloc[splits[si]:splits[si + 1]]
            section_buckets[si].extend(chunk.index.tolist())

    # Phase 3: assign time slots
    df_work["Session ID"] = None
    df_work["Date"] = None
    df_work["Time Slot"] = None
    df_work["Section"] = None
    df_work["Overflow"] = ""

    session_id = 1
    for si, bucket in enumerate(section_buckets):
        sec = sections[si]
        section_pres = schedulable.loc[schedulable.index.isin(bucket)].copy()
        # Re-sort by Theme within section for grouping
        section_pres = section_pres.sort_values(by="Theme")

        time_cursor = sec["start_dt"]
        section_end = sec["end_dt"]

        if mode == "oral":
            for theme, theme_df in section_pres.groupby("Theme", sort=False):
                for i in range(0, len(theme_df), int(max_presentations)):
                    group = theme_df.iloc[i:i + int(max_presentations)]
                    required_end = time_cursor + timedelta(minutes=int(slot_duration))
                    is_overflow = required_end > section_end
                    if is_overflow:
                        warnings.append(f"{sec['name']} ({sec['date']}) ran out of time during theme '{theme}'.")
                    for idx in group.index:
                        df_work.at[idx, "Session ID"] = session_id
                        df_work.at[idx, "Date"] = sec["date"].strftime("%Y-%m-%d")
                        df_work.at[idx, "Time Slot"] = time_cursor.strftime("%I:%M %p")
                        df_work.at[idx, "Section"] = sec["name"]
                        if is_overflow:
                            df_work.at[idx, "Overflow"] = "OVERFLOW"
                    time_cursor += timedelta(minutes=int(slot_duration))
                    session_id += 1
        else:  # poster -- all posters display simultaneously, just assign day
            time_label = f"{sec['start_dt'].strftime('%I:%M %p')} - {sec['end_dt'].strftime('%I:%M %p')}"
            for idx in section_pres.index:
                df_work.at[idx, "Session ID"] = session_id
                df_work.at[idx, "Date"] = sec["date"].strftime("%Y-%m-%d")
                df_work.at[idx, "Time Slot"] = time_label
                df_work.at[idx, "Section"] = sec["name"]
                session_id += 1

    # Check for section imbalance
    for si, sec in enumerate(sections):
        count = len(section_buckets[si])
        if count > target_per_section * 1.3 and target_per_section > 0:
            warnings.append(f"{sec['name']} has {count} presentations (target ~{int(target_per_section)}) due to availability constraints.")

    # Cleanup temp columns
    df_work.drop(columns=["_parsed", "_allowed", "_ctype"], inplace=True, errors="ignore")
    if not excluded_df.empty:
        excluded_df = excluded_df.drop(columns=["_parsed", "_allowed", "_ctype", "Session ID", "Date", "Time Slot", "Section"], errors="ignore")

    return df_work, excluded_df, warnings, session_id


def _assign_no_constraints(df_work, sections, mode, slot_duration, max_presentations):
    """Original algorithm when no constraints are present."""
    num_sections = len(sections)
    split_indices = np.linspace(0, len(df_work), num_sections + 1, dtype=int)
    section_dfs = [df_work.iloc[split_indices[i]:split_indices[i + 1]] for i in range(num_sections)]

    df_work["Session ID"] = None
    df_work["Date"] = None
    df_work["Time Slot"] = None
    df_work["Section"] = None
    df_work["Overflow"] = ""

    session_id = 1
    current_time = [s["start_dt"] for s in sections]
    overflow_warnings = []

    if mode == "oral":
        for si, section_df in enumerate(section_dfs):
            section_end = sections[si]["end_dt"]
            for theme, theme_df in section_df.groupby("Theme"):
                for i in range(0, len(theme_df), int(max_presentations)):
                    group = theme_df.iloc[i:i + int(max_presentations)]
                    required_end = current_time[si] + timedelta(minutes=int(slot_duration))
                    is_overflow = required_end > section_end
                    if is_overflow:
                        overflow_warnings.append(f"{sections[si]['name']} ({sections[si]['date']}) ran out of time during theme '{theme}'.")
                    for idx in group.index:
                        df_work.at[idx, "Session ID"] = session_id
                        df_work.at[idx, "Date"] = sections[si]["date"].strftime("%Y-%m-%d")
                        df_work.at[idx, "Time Slot"] = current_time[si].strftime("%I:%M %p")
                        df_work.at[idx, "Section"] = sections[si]["name"]
                        if is_overflow:
                            df_work.at[idx, "Overflow"] = "OVERFLOW"
                    current_time[si] += timedelta(minutes=int(slot_duration))
                    session_id += 1
    else:  # poster -- all posters display simultaneously, just assign day
        global_session_id = 1
        for si, section_df in enumerate(section_dfs):
            time_label = f"{sections[si]['start_dt'].strftime('%I:%M %p')} - {sections[si]['end_dt'].strftime('%I:%M %p')}"
            for i in range(len(section_df)):
                df_work.at[section_df.index[i], "Session ID"] = global_session_id
                df_work.at[section_df.index[i], "Date"] = sections[si]["date"].strftime("%Y-%m-%d")
                df_work.at[section_df.index[i], "Time Slot"] = time_label
                df_work.at[section_df.index[i], "Section"] = sections[si]["name"]
                global_session_id += 1
        session_id = global_session_id

    excluded_df = pd.DataFrame()
    return df_work, excluded_df, overflow_warnings, session_id


def build_xlsx(sections_data, final_df, columns, output_filename, excluded_df=None):
    """
    sections_data: list of dicts with keys: sheet_name, section_df
    final_df: master dataframe with all rows
    columns: list of column names to write
    Returns BytesIO object.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # Master sheet first
    ws_master = wb.create_sheet("Master")
    style_sheet(ws_master, columns)
    write_rows(ws_master, final_df, columns)

    # One sheet per section
    for sd in sections_data:
        ws = wb.create_sheet(sd["sheet_name"])
        style_sheet(ws, columns)
        write_rows(ws, sd["section_df"], columns)

    # Excluded sheet (if any)
    if excluded_df is not None and not excluded_df.empty:
        excl_cols = ["Theme", "Title", "Presenter(s)", "Faculty Mentor", "Availability Constraint"]
        excl_cols = [c for c in excl_cols if c in excluded_df.columns]
        ws_excl = wb.create_sheet("Excluded")
        style_sheet(ws_excl, excl_cols)
        write_rows(ws_excl, excluded_df.reset_index(drop=True), excl_cols)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ----------------------------------------------------------------
# LAYOUT -- Tabbed interface: Original + CSUSB MOTM Edition
# ----------------------------------------------------------------
tab_original, tab_motm = st.tabs(["Original", "CSUSB MOTM Edition"])

# ==================================================================
# ORIGINAL TAB -- exact replica of the original author's code
# ==================================================================
with tab_original:
    orig_col1, orig_col2 = st.columns([1, 2])

    orig_col1.markdown('<div class="column-spacing"></div>', unsafe_allow_html=True)
    orig_col2.markdown('<div class="column-spacing"></div>', unsafe_allow_html=True)

    with orig_col1:
        st.markdown("""
        **Welcome to the Conference Schedule Maker!**
        Please follow the instructions below to ensure smooth processing of your schedule:

        #### Required Columns in the Excel File:
        Your Excel file must include **exactly** the following columns, spelled as shown (case-sensitive) and located in the first row:
        - **Theme**
        - **Title**
        - **Presenter(s)**
        - **Faculty Mentor**

        #### Column Descriptions:
        - **Theme**: Category or department (e.g., Arts, Biology, Computer Science)
        - **Title**: Title of the presentation
        - **Presenter(s)**: The name(s) of the presenter(s), separated by commas if there are multiple presenters
        - **Faculty Mentor**: The name(s) of the faculty mentor(s)

        """)

    with orig_col2:
        orig_uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"], key="orig_upload")

        if orig_uploaded_file:
            orig_df = pd.read_excel(orig_uploaded_file)

            orig_required = ['Theme', 'Title', 'Presenter(s)', 'Faculty Mentor']
            orig_missing = [col for col in orig_required if col not in orig_df.columns]

            if orig_missing:
                st.error(f"Missing columns in uploaded file: {orig_missing}")
            else:
                orig_session_type = st.radio("Choose Session Type:", ["Oral Session Maker", "Poster Session Maker"], key="orig_session_type")

                # Oral Session Maker
                if orig_session_type == "Oral Session Maker":
                    orig_slot = st.number_input("Duration per presentation (minutes):",
                                                min_value=5, max_value=20, value=15, key="orig_slot")
                    orig_max = st.number_input("Max presentations per session:",
                                               min_value=3, max_value=10, value=4, key="orig_max")
                    orig_num = st.number_input("Number of sections(or no. of oral sessions throughout the day):",
                                               min_value=1, max_value=3, value=2, key="orig_num")

                    orig_sections = []
                    for i in range(orig_num):
                        st.subheader(f"Section {i+1} Time Range")
                        oc1, oc2 = st.columns(2)
                        with oc1:
                            start = st.time_input(f"Start time:", key=f"orig_start_{i}",
                                                  value=datetime.strptime("10:00 AM", "%I:%M %p").time() if i == 0
                                                  else datetime.strptime("2:00 PM", "%I:%M %p").time())
                        with oc2:
                            end = st.time_input(f"End time:", key=f"orig_end_{i}",
                                                value=datetime.strptime("11:00 AM", "%I:%M %p").time() if i == 0
                                                else datetime.strptime("3:00 PM", "%I:%M %p").time())

                        orig_sections.append({
                            'name': f"Section {i+1}",
                            'start': start,
                            'end': end,
                            'start_dt': datetime.combine(datetime.today(), start),
                            'end_dt': datetime.combine(
                                datetime.today() + timedelta(days=1) if end < start else datetime.today(),
                                end
                            )
                        })

                    if st.button("Generate Schedule", key="orig_gen_oral"):
                        orig_df = orig_df.sort_values(by="Theme")
                        total_presentations = len(orig_df)
                        split_indices = np.linspace(0, total_presentations, orig_num+1, dtype=int)
                        section_dfs = [orig_df.iloc[split_indices[i]:split_indices[i+1]] for i in range(orig_num)]

                        orig_df['Session ID'] = None
                        orig_df['Time Slot'] = None
                        orig_df['Section'] = None

                        session_id = 1
                        current_time = [section['start_dt'] for section in orig_sections]
                        section_day = [0] * orig_num

                        for section_idx, section_df in enumerate(section_dfs):
                            theme_groups = section_df.groupby('Theme')

                            for theme, theme_df in theme_groups:
                                for i in range(0, len(theme_df), orig_max):
                                    presentation_group = theme_df.iloc[i:i+orig_max]
                                    required_time = current_time[section_idx] + timedelta(minutes=len(presentation_group)*orig_slot)
                                    section_end = datetime.combine(
                                        datetime.today() + timedelta(days=section_day[section_idx]),
                                        orig_sections[section_idx]['end']
                                    )

                                    if required_time > section_end:
                                        section_day[section_idx] += 1
                                        current_time[section_idx] = datetime.combine(
                                            datetime.today() + timedelta(days=section_day[section_idx]),
                                            orig_sections[section_idx]['start']
                                        )
                                        required_time = current_time[section_idx] + timedelta(minutes=len(presentation_group)*orig_slot)

                                    time_cursor = current_time[section_idx]
                                    for idx in presentation_group.index:
                                        orig_df.at[idx, 'Session ID'] = session_id
                                        orig_df.at[idx, 'Time Slot'] = time_cursor.strftime("%I:%M %p")
                                        orig_df.at[idx, 'Section'] = orig_sections[section_idx]['name']
                                        time_cursor += timedelta(minutes=orig_slot)

                                    current_time[section_idx] = time_cursor
                                    session_id += 1

                        final_df = orig_df[['Section', 'Session ID', 'Time Slot', 'Theme', 'Title', 'Presenter(s)', 'Faculty Mentor']]
                        st.write("**Oral Schedule Preview:**")
                        st.dataframe(final_df.head(20))
                        scheduled = final_df[final_df['Session ID'].notna()]
                        st.write(f"Total scheduled: {len(scheduled)} presentations")
                        st.write(f"Total sessions created: {session_id-1}")

                        csv = final_df.to_csv(index=False).encode('utf-8')
                        st.download_button("Download Oral Schedule CSV", csv,
                                           "oral_presentation_schedule.csv", "text/csv",
                                           key="orig_dl_oral")

                # Poster Session Maker
                elif orig_session_type == "Poster Session Maker":
                    orig_poster_num = st.number_input("Number of poster sections(or no. of poster sessions throughout the day) :",
                                                      min_value=1, max_value=5, value=2, key="orig_poster_num")

                    orig_poster_sections = []
                    for i in range(orig_poster_num):
                        st.subheader(f"Poster Section {i+1} Time Range")
                        oc1, oc2 = st.columns(2)
                        with oc1:
                            start = st.time_input(f"Start Time:", key=f"orig_poster_start_{i}",
                                                  value=datetime.strptime("10:00 AM", "%I:%M %p").time() if i == 0
                                                  else datetime.strptime("1:00 PM", "%I:%M %p").time())
                        with oc2:
                            end = st.time_input(f"End Time:", key=f"orig_poster_end_{i}",
                                                value=datetime.strptime("11:30 AM", "%I:%M %p").time() if i == 0
                                                else datetime.strptime("2:30 PM", "%I:%M %p").time())
                        orig_poster_sections.append({
                            'name': f"Poster Section {i+1}",
                            'start': start,
                            'end': end,
                            'start_dt': datetime.combine(datetime.today(), start),
                            'end_dt': datetime.combine(
                                datetime.today() + timedelta(days=1) if end < start else datetime.today(),
                                end
                            )
                        })

                    if st.button("Generate Poster Schedule", key="orig_gen_poster"):
                        orig_df = orig_df.sort_values(by="Theme")
                        total_presentations = len(orig_df)
                        split_indices = np.linspace(0, total_presentations, orig_poster_num+1, dtype=int)
                        poster_groups = [orig_df.iloc[split_indices[i]:split_indices[i+1]] for i in range(orig_poster_num)]

                        orig_df['Session ID'] = None
                        orig_df['Time Slot'] = None
                        orig_df['Section'] = None

                        for idx, group in enumerate(poster_groups):
                            start_time = orig_poster_sections[idx]['start_dt']
                            end_time = orig_poster_sections[idx]['end_dt']
                            time_cursor = start_time
                            session_id = 1
                            for i in range(len(group)):
                                if time_cursor >= end_time:
                                    time_cursor = start_time + timedelta(days=1)

                                orig_df.at[group.index[i], 'Session ID'] = session_id
                                orig_df.at[group.index[i], 'Time Slot'] = time_cursor.strftime("%I:%M %p")
                                orig_df.at[group.index[i], 'Section'] = orig_poster_sections[idx]['name']
                                time_cursor += timedelta(minutes=10)
                                session_id += 1

                        final_df = orig_df[['Section', 'Session ID', 'Time Slot', 'Theme', 'Title', 'Presenter(s)', 'Faculty Mentor']]
                        st.write("**Poster Schedule Preview:**")
                        st.dataframe(final_df.head(20))
                        scheduled = final_df[final_df['Session ID'].notna()]
                        st.write(f"Total scheduled: {len(scheduled)} posters")
                        st.write(f"Total sessions created: {session_id-1}")

                        csv = final_df.to_csv(index=False).encode('utf-8')
                        st.download_button("Download Poster Schedule CSV", csv,
                                           "poster_presentation_schedule.csv", "text/csv",
                                           key="orig_dl_poster")

# ==================================================================
# CSUSB MOTM EDITION TAB
# ==================================================================
with tab_motm:
    motm_col1, motm_col2 = st.columns([1, 2])
    motm_col1.markdown('<div class="column-spacing"></div>', unsafe_allow_html=True)
    motm_col2.markdown('<div class="column-spacing"></div>', unsafe_allow_html=True)

    with motm_col1:
        st.markdown("""
        **CSUSB Meeting of the Minds Edition**

        This version adds availability-constraint-aware scheduling.

        #### Required Columns:
        - **Theme**
        - **Title**
        - **Presenter(s)**
        - **Faculty Mentor**

        #### Optional Column:
        - **Availability Constraint** -- supported formats:
          - *Empty / blank*: No constraint (schedule anywhere)
          - *Specific time window*: `April 15, 2:30 - 4:00 PM`
          - *Multiple windows*: comma-separated list of windows
          - *Day only*: `April 15 only`
          - *Either day*: No constraint
          - *None / Neither day*: Presenter excluded from schedule
          - *Late submission*: Treated as no constraint

        #### Poster Mode:
        Posters are assigned to a **day/section** only (no individual time slots),
        since all posters display simultaneously.

        #### Output:
        XLSX with **Master** + per-section tabs + **Excluded** sheet (if any).
        """)

    with motm_col2:
        motm_file = st.file_uploader("Upload Excel File", type=["xlsx"], key="motm_upload")

        if motm_file:
            motm_sheets = pd.read_excel(motm_file, sheet_name=None)
            if "Master" in motm_sheets:
                df = motm_sheets["Master"]
            else:
                df = list(motm_sheets.values())[0]

            required_columns = ['Theme', 'Title', 'Presenter(s)', 'Faculty Mentor']
            missing_cols = [c for c in required_columns if c not in df.columns]

            if missing_cols:
                st.error(f"Missing columns: {missing_cols}")
            elif df.empty:
                st.warning("The Master sheet is empty. Add your presentations to the Master sheet and re-upload.")
            else:
                has_constraints = "Availability Constraint" in df.columns

                if has_constraints:
                    _year = datetime.today().year
                    _parsed = df["Availability Constraint"].apply(lambda v: parse_constraint(v, _year))
                    _n_any = sum(1 for p in _parsed if p["type"] == "any")
                    _n_excl = sum(1 for p in _parsed if p["type"] == "excluded")
                    _n_unrec = sum(1 for p in _parsed if p["type"] == "unrecognized")
                    _n_constrained = len(df) - _n_any - _n_excl - _n_unrec
                    with st.expander("Availability Constraint Summary", expanded=False):
                        st.write(f"Total presentations: **{len(df)}**")
                        st.write(f"Unconstrained: **{_n_any}**")
                        st.write(f"Constrained to specific sections: **{_n_constrained}**")
                        st.write(f"Excluded (cannot attend): **{_n_excl}**")
                        if _n_unrec > 0:
                            st.warning(f"{_n_unrec} presentation(s) have unrecognized constraint formats (will be treated as unconstrained).")
                        if _n_excl > 0:
                            _excl_df = df[_parsed.apply(lambda p: p["type"] == "excluded")]
                            st.info("Excluded presenters:")
                            st.dataframe(_excl_df[["Title", "Presenter(s)", "Availability Constraint"]])

                session_type = st.radio("Choose Session Type:", ["Oral Session Maker", "Poster Session Maker"], key="motm_session_type")

                if session_type == "Oral Session Maker":
                    slot_duration = st.number_input("Duration per presentation (minutes):", min_value=5, max_value=20, value=15, key="motm_slot")
                    max_presentations = st.number_input("Max presentations per session:", min_value=3, max_value=10, value=4, key="motm_max")
                    num_sections = st.number_input("Number of sections (across all days):", min_value=1, max_value=10, value=2, key="motm_num")

                    sections = []
                    for i in range(int(num_sections)):
                        st.subheader(f"Section {i+1}")
                        c_date, c_start, c_end = st.columns(3)
                        with c_date:
                            section_date = st.date_input("Date:", key=f"motm_date_{i}", value=datetime.today().date())
                        with c_start:
                            start = st.time_input("Start time:", key=f"motm_start_{i}",
                                                  value=datetime.strptime("10:00 AM", "%I:%M %p").time() if i == 0
                                                  else datetime.strptime("2:00 PM", "%I:%M %p").time())
                        with c_end:
                            end = st.time_input("End time:", key=f"motm_end_{i}",
                                                value=datetime.strptime("11:00 AM", "%I:%M %p").time() if i == 0
                                                else datetime.strptime("3:00 PM", "%I:%M %p").time())

                        start_dt = datetime.combine(section_date, start)
                        end_dt = datetime.combine(section_date, end)
                        if end_dt <= start_dt:
                            end_dt += timedelta(days=1)

                        sections.append({
                            'name': f"Section {i+1}", 'sheet_name': make_sheet_name(section_date, start, end),
                            'date': section_date, 'start': start, 'end': end,
                            'start_dt': start_dt, 'end_dt': end_dt,
                        })

                    if st.button("Generate Schedule", key="motm_gen_oral"):
                        df_work = df.copy().sort_values(by="Theme").reset_index(drop=True)

                        df_work, excluded_df, gen_warnings, session_id = assign_with_constraints(
                            df_work, sections, has_constraints, mode="oral",
                            slot_duration=slot_duration, max_presentations=max_presentations,
                        )

                        for w in gen_warnings:
                            st.warning(w)

                        cols = ['Theme', 'Title', 'Presenter(s)', 'Faculty Mentor']
                        if has_constraints:
                            cols.append('Availability Constraint')
                        cols += ['Section', 'Date', 'Session ID', 'Time Slot', 'Overflow']

                        final_df = df_work[df_work['Section'].notna()][cols].sort_values(by=['Section', 'Session ID']).reset_index(drop=True)

                        overflow_count = (final_df['Overflow'] == 'OVERFLOW').sum()
                        st.write("**Oral Schedule Preview (first 20 rows):**")
                        st.dataframe(final_df.head(20))
                        if len(final_df) > 20:
                            st.caption(f"Showing 20 of {len(final_df)} rows.")
                        st.write(f"Total scheduled: {final_df['Session ID'].notna().sum()} presentations | Sessions: {session_id-1}")
                        if overflow_count > 0:
                            st.warning(f"{overflow_count} presentation(s) tagged as OVERFLOW (scheduled past section end time).")
                        if not excluded_df.empty:
                            st.info(f"{len(excluded_df)} presenter(s) excluded due to availability constraints (see Excluded sheet in download).")

                        sections_data = []
                        for si, s in enumerate(sections):
                            mask = final_df['Section'] == s['name']
                            sections_data.append({"sheet_name": s['sheet_name'], "section_df": final_df[mask].reset_index(drop=True)})

                        xlsx_buf = build_xlsx(sections_data, final_df, cols, "oral_schedule.xlsx",
                                              excluded_df=excluded_df if not excluded_df.empty else None)
                        st.download_button("Download Oral Schedule XLSX", xlsx_buf,
                                           "oral_presentation_schedule.xlsx",
                                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                           key="motm_dl_oral")

                elif session_type == "Poster Session Maker":
                    num_sections = st.number_input("Number of poster sections (across all days):", min_value=1, max_value=10, value=2, key="motm_poster_num")

                    poster_sections = []
                    for i in range(int(num_sections)):
                        st.subheader(f"Poster Section {i+1}")
                        c_date, c_start, c_end = st.columns(3)
                        with c_date:
                            section_date = st.date_input("Date:", key=f"motm_poster_date_{i}", value=datetime.today().date())
                        with c_start:
                            start = st.time_input("Start Time:", key=f"motm_poster_start_{i}",
                                                  value=datetime.strptime("10:00 AM", "%I:%M %p").time() if i == 0
                                                  else datetime.strptime("1:00 PM", "%I:%M %p").time())
                        with c_end:
                            end = st.time_input("End Time:", key=f"motm_poster_end_{i}",
                                                value=datetime.strptime("11:30 AM", "%I:%M %p").time() if i == 0
                                                else datetime.strptime("2:30 PM", "%I:%M %p").time())

                        start_dt = datetime.combine(section_date, start)
                        end_dt = datetime.combine(section_date, end)
                        if end_dt <= start_dt:
                            end_dt += timedelta(days=1)

                        poster_sections.append({
                            'name': f"Poster Section {i+1}", 'sheet_name': make_sheet_name(section_date, start, end),
                            'date': section_date, 'start': start, 'end': end,
                            'start_dt': start_dt, 'end_dt': end_dt,
                        })

                    if st.button("Generate Poster Schedule", key="motm_gen_poster"):
                        df_work = df.copy().sort_values(by="Theme").reset_index(drop=True)

                        df_work, excluded_df, gen_warnings, session_id = assign_with_constraints(
                            df_work, poster_sections, has_constraints, mode="poster",
                        )

                        for w in gen_warnings:
                            st.warning(w)

                        cols = ['Theme', 'Title', 'Presenter(s)', 'Faculty Mentor']
                        if has_constraints:
                            cols.append('Availability Constraint')
                        cols += ['Section', 'Date', 'Session ID', 'Time Slot']

                        final_df = df_work[df_work['Section'].notna()][cols].sort_values(by=['Section', 'Session ID']).reset_index(drop=True)

                        st.write("**Poster Schedule Preview (first 20 rows):**")
                        st.dataframe(final_df.head(20))
                        if len(final_df) > 20:
                            st.caption(f"Showing 20 of {len(final_df)} rows.")
                        st.write(f"Total scheduled: {final_df['Session ID'].notna().sum()} posters | Slots: {session_id-1}")
                        if not excluded_df.empty:
                            st.info(f"{len(excluded_df)} presenter(s) excluded due to availability constraints (see Excluded sheet in download).")

                        sections_data = []
                        for si, s in enumerate(poster_sections):
                            mask = final_df['Section'] == s['name']
                            sections_data.append({"sheet_name": s['sheet_name'], "section_df": final_df[mask].reset_index(drop=True)})

                        xlsx_buf = build_xlsx(sections_data, final_df, cols, "poster_schedule.xlsx",
                                              excluded_df=excluded_df if not excluded_df.empty else None)
                        st.download_button("Download Poster Schedule XLSX", xlsx_buf,
                                           "poster_presentation_schedule.xlsx",
                                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                           key="motm_dl_poster")
